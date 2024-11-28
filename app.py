import os
import random
import time
from openpyxl import Workbook, load_workbook
from flask import Flask, render_template, request, redirect, url_for
from apscheduler.schedulers.background import BackgroundScheduler

app = Flask(__name__)

# Paths to Excel files
STUDENT_DATA_FILE = "student_data.xlsx"
WINNER_FILE = "LuckyWinner.xlsx"

# Create the winner file if it doesn't exist
if not os.path.exists(WINNER_FILE):
    wb = Workbook()
    ws = wb.active
    ws.title = "Winners"
    ws.append(["Student Name", "College Name", "Contact No"])  # Add headers
    wb.save(WINNER_FILE)

# Initialize round number
round_number = 1
current_winner = None
winner_selected = False

# Winner selection function
def select_winner():
    global round_number, current_winner, winner_selected

    # Load student data
    if os.path.exists(STUDENT_DATA_FILE):
        wb = load_workbook(STUDENT_DATA_FILE)
        ws = wb.active

        # Select a random winner
        students = list(ws.iter_rows(values_only=True))
        if len(students) > 1:  # Ensure there are students to select from
            winner = random.choice(students[1:])  # Skip header row
            current_winner = winner[:2]  # Exclude contact number
            winner_selected = True

            # Save the winner to the winner file
            winner_wb = load_workbook(WINNER_FILE)
            winner_ws = winner_wb.active
            winner_ws.append(winner)
            winner_wb.save(WINNER_FILE)

            # Create a new round file
            round_file = f"round{round_number}.xlsx"
            round_wb = Workbook()
            round_ws = round_wb.active
            round_ws.title = "Winners"
            round_ws.append(["Student Name", "College Name", "Contact No"])  # Add headers

            # Copy all student data to the new round file, starting from the second row
            for row in students[1:]:
                round_ws.append(row)
            round_wb.save(round_file)

            # Increment round number
            round_number += 1

            # Clear student data
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                for cell in row:
                    cell.value = None
            wb.save(STUDENT_DATA_FILE)

@app.route("/")
def index():
    global winner_selected, current_winner
    winner_data = None
    if winner_selected:
        winner_data = current_winner
        winner_selected = False
    return render_template("form.html", winner=winner_data)

@app.route("/submit", methods=["POST"])
def submit():
    # Get data from the form
    student_name = request.form["student_name"]
    college_name = request.form["college_name"]
    contact_no = request.form["contact_no"]

    # Check if the Excel file exists
    if not os.path.exists(STUDENT_DATA_FILE):
        wb = Workbook()
        ws = wb.active
        ws.title = "Students"
        ws.append(["Student Name", "College Name", "Contact No"])  # Add headers
    else:
        wb = load_workbook(STUDENT_DATA_FILE)
        ws = wb.active

    # Append the new student data
    ws.append([student_name, college_name, contact_no])
    wb.save(STUDENT_DATA_FILE)

    return redirect(url_for('index'))

if __name__ == '__main__':
    # Start the winner selection using APScheduler
    scheduler = BackgroundScheduler()
    scheduler.add_job(select_winner, 'interval', seconds=45)  # Change the interval here
    scheduler.start()

    app.run(debug=True)
